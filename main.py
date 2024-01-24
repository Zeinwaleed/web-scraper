from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time
driver = webdriver.Chrome()

driver.get("https://www.merckgroup.com/en/careers/job-search.html?")

jobs = driver.find_elements(By.CLASS_NAME ,"se-list-job-item-title")

joblist = []
for job in range(0,len(jobs) - 1):

    jobname = jobs[job].find_elements(By.TAG_NAME,"a")

    for each in jobname:
        
        joblist.append(each.get_attribute("href"))

jobnames = []
joblocations = []
for job in joblist:
    jobdriver = webdriver.Chrome()
    jobdriver.get(job)
    jobholder = jobdriver.find_elements(By.CLASS_NAME,"jd-header-content")

    if jobholder == []:
        continue
        print("page not found")

    
    joblocation = jobholder[0].find_elements(By.CLASS_NAME,"jd-header-text")[1]
    joblocations.append(joblocation.text)
    jobname = jobholder[0].find_element(By.CSS_SELECTOR,".h2.jd-header-title")
    jobnames.append(jobname.text)






wb = load_workbook("test.xlsx")

ws = wb.active
i = 1

while i < len(jobnames):
    ws[f"A{i + 1}"].value = jobnames[i]
    ws[f"B{i + 1}"].value = joblocations[i]
    
    i += 1

# Save the workbook after the loop
wb.save("test.xlsx")